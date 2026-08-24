from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from matcher import DAY_ORDER, Observation, WeeklyData
from members import Member
from storage import AliasStore


HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
IMPORTED_FONT = "008000"
STATIC_FONT = "666666"
REVIEW_FILL = "FCE4D6"
ERROR_FILL = "F4CCCC"


def _style_header(ws, row: int, max_col: int) -> None:
    for cell in ws[row][:max_col]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center")


def _autowidth(ws, min_width: int = 10, max_width: int = 44) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)) + 2)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_width, min(max_width, width))


def export_weekly_workbook(
    path: Path,
    members: list[Member],
    weekly: WeeklyData,
    alias_store: AliasStore,
    member_source: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Scores"
    ws.sheet_view.showGridLines = False

    headers = ["ID", "Player"] + [day.title() for day in DAY_ORDER]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "C2"

    for member in sorted(members, key=lambda m: m.member_id):
        row = [member.member_id, member.name]
        for day in DAY_ORDER:
            row.append(weekly.scores.get((member.member_id, day)))
        ws.append(row)
        ws.cell(ws.max_row, 1).font = Font(color=STATIC_FONT)
        ws.cell(ws.max_row, 2).font = Font(color=STATIC_FONT)
        for c in range(3, 3 + len(DAY_ORDER)):
            cell = ws.cell(ws.max_row, c)
            if cell.value is not None:
                cell.number_format = "#,##0"
                cell.font = Font(color=IMPORTED_FONT)

    _autowidth(ws)

    obs_ws = wb.create_sheet("Observations")
    obs_ws.sheet_view.showGridLines = False
    obs_headers = [
        "Source file", "Day", "Rank", "Visible player ID", "Raw name", "Points",
        "Extraction confidence", "Pinned row", "Matched member ID", "Matched member",
        "Match method", "Match confidence", "Issue"
    ]
    obs_ws.append(obs_headers)
    _style_header(obs_ws, 1, len(obs_headers))
    obs_ws.freeze_panes = "A2"
    for obs in weekly.observations:
        obs_ws.append([
            obs.source_file, obs.day, obs.rank, obs.raw_player_id, obs.raw_name, obs.points,
            obs.extraction_confidence, obs.is_pinned_row, obs.matched_member_id,
            obs.matched_member_name, obs.match_method, obs.match_confidence, obs.issue,
        ])
        r = obs_ws.max_row
        obs_ws.cell(r, 6).number_format = "#,##0"
        obs_ws.cell(r, 7).number_format = "0%"
        obs_ws.cell(r, 12).number_format = "0%"
        if obs.issue:
            for cell in obs_ws[r]:
                cell.fill = PatternFill("solid", fgColor=REVIEW_FILL)
    _autowidth(obs_ws)

    issues_ws = wb.create_sheet("Issues")
    issues_ws.sheet_view.showGridLines = False
    issues_ws.append(["Type", "Details"])
    _style_header(issues_ws, 1, 2)
    for issue in weekly.issues:
        issues_ws.append(["Review", issue])
        for cell in issues_ws[issues_ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=REVIEW_FILL)
    for day, missing in weekly.missing_by_day.items():
        for member in missing:
            issues_ws.append(["Missing player", f"{day.title()}: {member.member_id} - {member.name}"])
    _autowidth(issues_ws, max_width=90)

    alias_ws = wb.create_sheet("Aliases")
    alias_ws.sheet_view.showGridLines = False
    alias_ws.append(["Observed alias", "Member ID"])
    _style_header(alias_ws, 1, 2)
    for alias, member_id in alias_store.list_aliases():
        alias_ws.append([alias, member_id])
    _autowidth(alias_ws)

    meta_ws = wb.create_sheet("Run Info")
    meta_ws.sheet_view.showGridLines = False
    meta_ws.append(["Setting", "Value"])
    _style_header(meta_ws, 1, 2)
    meta_ws.append(["Members source", member_source])
    meta_ws.append(["Active members", len(members)])
    meta_ws.append(["Observations", len(weekly.observations)])
    meta_ws.append(["Issues", len(weekly.issues)])
    _autowidth(meta_ws, max_width=100)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
