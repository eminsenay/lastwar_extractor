from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace

from openpyxl import Workbook

from matcher import MemberMatcher, build_weekly_data, observations_from_extractions
from members import Member, load_members_from_google_sheet, load_members_from_xlsx
from avatars import AvatarStore, fingerprint_from_bbox
from storage import AliasStore
from excel_export import export_weekly_workbook


def row(rank, player_id, raw_name, points, confidence=.99):
    return SimpleNamespace(
        rank=rank, player_id=player_id, raw_name=raw_name,
        points=points, extraction_confidence=confidence
    )


def test_google_sheet_download_cleanup():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'
    ws.append(['ID', 'Name', 'Rank'])
    ws.append([101, 'Alpha Player', 'Member'])
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)
    data = payload.getvalue()
    wb.close()

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return data

    import members as members_module
    original = members_module.urllib.request.urlopen
    members_module.urllib.request.urlopen = lambda req, timeout=30: FakeResponse()
    try:
        result = load_members_from_google_sheet(
            'https://docs.google.com/spreadsheets/d/abc123/edit#gid=0',
            sheet_name='Members',
            timeout=30,
        )
        assert result.members[0].member_id == 101
        assert result.source_description == 'https://docs.google.com/spreadsheets/d/abc123/edit#gid=0'
    finally:
        members_module.urllib.request.urlopen = original



def test_avatar_assisted_matching():
    from PIL import Image, ImageDraw

    with TemporaryDirectory() as td:
        td_path = Path(td)
        # Create two screenshots with the same feature-rich central avatar but
        # different outer frames / image sizes.
        base = Image.new("RGB", (120, 120), "white")
        draw = ImageDraw.Draw(base)
        for i in range(0, 120, 12):
            draw.line((0, i, 119, 119 - i), fill=(20 + i, 80, 160), width=3)
            draw.rectangle((i // 2, i // 3, 30 + i // 2, 25 + i // 3), outline=(180, 30, 30), width=2)
        draw.ellipse((28, 20, 92, 84), outline=(10, 10, 10), width=5)
        draw.line((45, 55, 75, 55), fill=(10, 10, 10), width=4)

        shot1 = Image.new("RGB", (300, 300), (210, 220, 235))
        shot1.paste(base, (90, 90))
        shot1_path = td_path / "one.png"
        shot1.save(shot1_path)

        shot2 = Image.new("RGB", (420, 420), (230, 220, 210))
        framed = base.resize((150, 150))
        shot2.paste(framed, (135, 135))
        shot2_path = td_path / "two.png"
        shot2.save(shot2_path)

        fp1 = fingerprint_from_bbox(shot1_path, (300, 300, 400, 400))
        fp2 = fingerprint_from_bbox(shot2_path, (321, 321, 357, 357))
        assert fp1 and fp2

        members = [
            Member(1, "Alpha", "R1", None, None),
            Member(2, "Beta", "R1", None, None),
        ]
        alias_store = AliasStore(td_path / "app.sqlite3")
        avatar_store = AvatarStore(td_path / "app.sqlite3")
        matcher = MemberMatcher(members, alias_store, avatar_store)

        trusted = row(1, 1, "Alpha", 100)
        trusted.avatar_bbox = None
        from matcher import Observation
        obs1 = Observation("one.png", "tuesday", 1, 1, "Alpha", 100, .99, False,
                           avatar_fingerprint=fp1)
        matcher.match_deterministic(obs1)
        assert obs1.matched_member_id == 1
        assert matcher.learn_avatar(obs1)

        renamed = Observation("two.png", "saturday", 5, None, "CompletelyNewName", 200, .99, False,
                              avatar_fingerprint=fp2)
        matcher.match_deterministic(renamed)
        assert renamed.matched_member_id is None
        matcher.match_avatar(renamed)
        assert renamed.matched_member_id == 1
        assert renamed.match_method == "avatar_auto"


def main():
    test_avatar_assisted_matching()
    test_google_sheet_download_cleanup()
    workbook = Path('/mnt/data/LastWar-1537-EfC-R4.xlsx')
    loaded = load_members_from_xlsx(workbook)
    assert len(loaded.members) == 100
    assert all(m.rank.casefold() != 'left' for m in loaded.members)
    assert any('184' in w for w in loaded.warnings)

    extraction = SimpleNamespace(
        detected_day='thursday',
        rows=[
            row(1, 164, 'Janninus', 360962392),
            row(2, 41, 'Mookha', 305798489),
            row(3, None, 'UnknownRenamedPlayer', 123456, .80),
        ],
        pinned_row=row(42, None, 'KingOfGondor', 77844315),
        warnings=[]
    )
    result = SimpleNamespace(image_path=Path('sample.png'), extraction=extraction, error=None)

    with TemporaryDirectory() as td:
        store = AliasStore(Path(td) / 'aliases.sqlite3')
        matcher = MemberMatcher(loaded.members, store)
        obs, issues = observations_from_extractions([result])
        for o in obs:
            matcher.match(o)
        assert obs[0].matched_member_id == 164 and obs[0].match_method == 'id'
        assert obs[1].matched_member_id == 41 and obs[1].match_method == 'id'
        assert obs[2].matched_member_id is None and obs[2].alternatives
        assert obs[3].matched_member_id == 5 and obs[3].match_method == 'exact_name'

        matcher.manual_assign(obs[2], 3, remember_alias=True)
        assert store.get_member_id('unknownrenamedplayer') == 3
        weekly = build_weekly_data(obs, loaded.members, issues)
        assert weekly.scores[(164, 'thursday')] == 360962392
        assert weekly.scores[(5, 'thursday')] == 77844315
        out = Path(td) / 'weekly.xlsx'
        export_weekly_workbook(out, loaded.members, weekly, store, str(workbook))
        assert out.exists() and out.stat().st_size > 0

        # Validate produced workbook structure.
        from openpyxl import load_workbook
        wb = load_workbook(out, data_only=False)
        assert wb.sheetnames == ['Weekly Scores', 'Observations', 'Issues', 'Aliases', 'Run Info']
        ws = wb['Weekly Scores']
        assert [ws.cell(1, c).value for c in range(1, 9)] == [
            'ID', 'Player', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'
        ]
        # Find Janninus and verify Thursday score.
        found = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == 164:
                assert ws.cell(r, 6).value == 360962392
                found = True
        assert found

    print('offline tests passed')


if __name__ == '__main__':
    main()
