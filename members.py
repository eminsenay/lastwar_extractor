from __future__ import annotations

import re
import tempfile
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class Member:
    member_id: int
    name: str
    rank: str
    joined_on: datetime | None
    total_hero_power: float | None


@dataclass(slots=True)
class MemberLoadResult:
    members: list[Member]
    all_rows: list[Member]
    warnings: list[str]
    source_description: str


def _coerce_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def load_members_from_xlsx(path: Path, sheet_name: str = "Members") -> MemberLoadResult:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Worksheet {sheet_name!r} not found. Available: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError("Members worksheet is empty")

        index = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
        required = ["ID", "Name", "Rank"]
        missing = [name for name in required if name not in index]
        if missing:
            raise ValueError(f"Members worksheet is missing columns: {', '.join(missing)}")

        all_members: list[Member] = []
        warnings: list[str] = []
        for excel_row, row in enumerate(rows, start=2):
            member_id = _coerce_int(row[index["ID"]] if index["ID"] < len(row) else None)
            name = row[index["Name"]] if index["Name"] < len(row) else None
            rank = row[index["Rank"]] if index["Rank"] < len(row) else None
            if member_id is None and not name:
                continue
            if member_id is None or not name:
                warnings.append(f"Members row {excel_row} has missing/invalid ID or name and was skipped.")
                continue
            joined = row[index.get("Joined on", -1)] if index.get("Joined on", -1) >= 0 and index["Joined on"] < len(row) else None
            hero_power = row[index.get("Total Hero Power", -1)] if index.get("Total Hero Power", -1) >= 0 and index["Total Hero Power"] < len(row) else None
            try:
                hero_power = float(hero_power) if hero_power is not None else None
            except (TypeError, ValueError):
                hero_power = None
            all_members.append(Member(member_id, str(name).strip(), str(rank or "").strip(), joined, hero_power))

        active = [m for m in all_members if m.rank.casefold() != "left"]

        by_id: dict[int, list[Member]] = {}
        for m in active:
            by_id.setdefault(m.member_id, []).append(m)
        for member_id, group in by_id.items():
            if len(group) > 1:
                warnings.append(
                    f"Duplicate active member ID {member_id}: " + ", ".join(m.name for m in group)
                )

        historical_by_id: dict[int, list[Member]] = {}
        for m in all_members:
            historical_by_id.setdefault(m.member_id, []).append(m)
        for member_id, group in historical_by_id.items():
            if len(group) > 1 and member_id not in {i for i, g in by_id.items() if len(g) > 1}:
                warnings.append(
                    f"Member ID {member_id} occurs multiple times historically; active filtering resolves it: "
                    + ", ".join(f"{m.name} ({m.rank})" for m in group)
                )

        return MemberLoadResult(active, all_members, warnings, str(path))
    finally:
        wb.close()


def _google_sheet_export_url(url: str) -> str:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        raise ValueError("Could not find a Google spreadsheet ID in the URL")
    spreadsheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"


def load_members_from_google_sheet(url: str, sheet_name: str = "Members", timeout: int = 30) -> MemberLoadResult:
    export_url = _google_sheet_export_url(url)
    req = urllib.request.Request(export_url, headers={"User-Agent": "LastWarWeeklyApp/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            data = response.read()
    except Exception as exc:
        raise RuntimeError(
            "Could not download the Google Sheet anonymously. If your organization blocks this, "
            "use the local Excel workbook source instead."
        ) from exc

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)

    try:
        result = load_members_from_xlsx(tmp_path, sheet_name=sheet_name)
        result.source_description = url
        return result
    finally:
        tmp_path.unlink(missing_ok=True)
